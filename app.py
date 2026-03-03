import os
import io
import json
import threading
from datetime import datetime

# Candado global — evita que dos usuarios escriban al mismo tiempo
write_lock = threading.Lock()
from flask import Flask, request, jsonify, session, send_from_directory
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.files.file import File
from openpyxl import load_workbook

app = Flask(__name__, static_folder='public', static_url_path='')

# ═══════════════════════════════════════════
#  🔐 CREDENCIALES — solo estas van en Render
# ═══════════════════════════════════════════
APP_USER    = os.environ.get('APP_USER',    'Usuario123')
APP_PASS    = os.environ.get('APP_PASS',    'Contraseña2026')
SECRET_KEY  = os.environ.get('SESSION_SECRET', 'crg-sodexo-2026')
SP_USUARIO  = os.environ.get('SP_USUARIO')   # tu email @sodexo.com
SP_PASSWORD = os.environ.get('SP_PASSWORD')  # tu contraseña de Sodexo

# ═══════════════════════════════════════════
#  🌐 SHAREPOINT — rutas fijas (no son secretas)
# ═══════════════════════════════════════════
SITE_COMMAND = "https://sodexo.sharepoint.com/sites/Basededatos-CommandCenter"

RUTA_BD_CECO    = "/sites/Basededatos-CommandCenter/Documents partages/General/Banca/Automatizaciones/Codigos/Hub Tickets/Bd_nomenclaturas/Bd Ceco.xlsx"
RUTA_BD_ESTADO  = "/sites/Basededatos-CommandCenter/Documents partages/General/Banca/Automatizaciones/Codigos/Hub Tickets/Bd_nomenclaturas/Bd Estado cliente.xlsx"
RUTA_BD_FAMILIA = "/sites/Basededatos-CommandCenter/Documents partages/General/Banca/Automatizaciones/Codigos/Hub Tickets/Bd_nomenclaturas/Bd Familia y Sub familia.xlsx"
RUTA_BD_RESP    = "/sites/Basededatos-CommandCenter/Documents partages/General/Banca/Automatizaciones/Codigos/Hub Tickets/Bd_nomenclaturas/Bd Responsable.xlsx"

SITE_TICKETS    = "https://sodexo.sharepoint.com/sites/Basededatos-CommandCenter"
RUTA_BD_TICKETS = "/sites/Basededatos-CommandCenter/Documents partages/General/Banca/Base de Datos/BD Temis/Registro web de tickets.xlsx"
SHEET_TICKETS   = "Hoja1"

app.secret_key = SECRET_KEY

# ═══════════════════════════════════════════
#  HELPER — leer Excel desde SharePoint
# ═══════════════════════════════════════════
def get_ctx(site_url):
    """Crea contexto de SharePoint con credenciales de Sodexo."""
    return ClientContext(site_url).with_credentials(
        UserCredential(SP_USUARIO, SP_PASSWORD)
    )

def read_excel_from_sp(site_url, file_path, sheet=0, header=0):
    """Lee un Excel de SharePoint y devuelve lista de filas [[col1, col2, ...]]."""
    ctx = get_ctx(site_url)
    response = File.open_binary(ctx, file_path)
    buffer = io.BytesIO(response.content)
    import openpyxl
    wb = openpyxl.load_workbook(buffer, data_only=True)
    
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]
    
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows

def write_row_to_sp(site_url, file_path, sheet_name, row_data):
    """Agrega una fila al final del Excel en SharePoint."""
    ctx = get_ctx(site_url)
    
    # Descargar el archivo
    response = File.open_binary(ctx, file_path)
    buffer = io.BytesIO(response.content)
    
    # Abrir con openpyxl
    wb = load_workbook(buffer)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
    
    # Agregar la fila al final
    ws.append(row_data)
    next_row = ws.max_row
    
    # Guardar en buffer
    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    
    # Subir de vuelta a SharePoint
    folder_path = '/'.join(file_path.split('/')[:-1])
    file_name   = file_path.split('/')[-1]
    
    folder = ctx.web.get_folder_by_server_relative_url(folder_path)
    folder.upload_file(file_name, out_buffer.getvalue()).execute_query()
    
    return next_row

# ═══════════════════════════════════════════
#  RUTAS
# ═══════════════════════════════════════════

@app.route('/')
def index():
    return send_from_directory('public', 'index.html')

# ── Login ──
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    if data.get('usuario') == APP_USER and data.get('password') == APP_PASS:
        session['authed'] = True
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'Usuario o contraseña incorrectos'}), 401

# ── Logout ──
@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'ok': True})

# ── Check sesión ──
@app.route('/api/check')
def check():
    return jsonify({'authed': session.get('authed', False)})

# ── Cargar datos de los desplegables ──
@app.route('/api/data')
def get_data():
    if not session.get('authed'):
        return jsonify({'error': 'No autorizado'}), 401

    try:
        # ── Ceco: columnas A=Cliente, B=Ceco ──
        ceco = {}
        rows_ceco = read_excel_from_sp(SITE_COMMAND, RUTA_BD_CECO)
        for row in rows_ceco[1:]:  # saltar encabezado
            cliente = str(row[0] or '').strip()
            valor   = str(row[1] or '').strip()
            if cliente and valor:
                if cliente not in ceco:
                    ceco[cliente] = []
                ceco[cliente].append(valor)

        # ── Estado: columnas A=Cliente, B=Estado ──
        estado = {}
        rows_est = read_excel_from_sp(SITE_COMMAND, RUTA_BD_ESTADO)
        for row in rows_est[1:]:
            cliente = str(row[0] or '').strip()
            valor   = str(row[1] or '').strip()
            if cliente and valor:
                if cliente not in estado:
                    estado[cliente] = []
                estado[cliente].append(valor)

        # ── Familia: columnas A=Cliente, B=Familia, C=Sub Familia ──
        familia = {}
        rows_fam = read_excel_from_sp(SITE_COMMAND, RUTA_BD_FAMILIA)
        for row in rows_fam[1:]:
            cliente = str(row[0] or '').strip()
            fam     = str(row[1] or '').strip()
            subfam  = str(row[2] or '').strip() if len(row) > 2 else ''
            if cliente and fam:
                if cliente not in familia:
                    familia[cliente] = {}
                if fam not in familia[cliente]:
                    familia[cliente][fam] = []
                if subfam:
                    familia[cliente][fam].append(subfam)

        # ── Responsable: columnas A=Nombre, B=Correo ──
        responsable = []
        rows_resp = read_excel_from_sp(SITE_COMMAND, RUTA_BD_RESP)
        for row in rows_resp[1:]:
            nombre = str(row[0] or '').strip()
            correo = str(row[1] or '').strip() if len(row) > 1 else ''
            if nombre:
                responsable.append({'nombre': nombre, 'correo': correo})

        return jsonify({
            'ceco':        ceco,
            'estado':      estado,
            'familia':     familia,
            'responsable': responsable
        })

    except Exception as e:
        print(f'Error cargando datos: {e}')
        return jsonify({'error': str(e)}), 500

# ── Guardar ticket ──
@app.route('/api/ticket', methods=['POST'])
def save_ticket():
    if not session.get('authed'):
        return jsonify({'error': 'No autorizado'}), 401

    try:
        d = request.get_json()

        now = datetime.now()

        def fmt_date(date_str):
            if not date_str:
                return ''
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
                return dt.strftime('%d/%m/%Y 00:00:00')
            except:
                return date_str

        def fmt_datetime(dt):
            return dt.strftime('%d/%m/%Y %H:%M:%S')

        # Orden exacto de columnas según la tabla entregada:
        # WO | Ceco | Familia | Sub Familia | Descripción OT | Usuario creó OT |
        # Correo | Detalle | Tipo OT | Tipo Ticket | Estado |
        # Fecha modif estado | Modificado por | Fecha Apertura |
        # Fecha Inicio Real | Prioridad | Cliente
        row = [
            d.get('wo', ''),
            d.get('ceco', ''),
            d.get('familia', ''),
            d.get('sub_familia', ''),
            d.get('descripcion_ot', ''),
            d.get('usuario', ''),
            d.get('correo', ''),
            d.get('detalle', ''),
            d.get('tipo_ot', ''),
            'RM',                               # Tipo de Ticket (predeterminado)
            'ACK',                              # Estado (predeterminado)
            fmt_date(d.get('fecha_apertura')),  # Fecha modif estado = Fecha apertura
            d.get('usuario', ''),               # Modificado por = Usuario que creó OT
            fmt_date(d.get('fecha_apertura')),  # Fecha de Apertura Cliente
            fmt_datetime(now),                  # Fecha de Inicio Real = ahora
            d.get('prioridad', ''),
            d.get('cliente', '')
        ]

        # El candado garantiza que solo un usuario escribe a la vez
        with write_lock:
            next_row = write_row_to_sp(
                SITE_TICKETS,
                RUTA_BD_TICKETS,
                SHEET_TICKETS,
                row
            )

        return jsonify({'ok': True, 'fila': next_row})

    except Exception as e:
        print(f'Error guardando ticket: {e}')
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 3000))
    app.run(host='0.0.0.0', port=port)
